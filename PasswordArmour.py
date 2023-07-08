import random
import string
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

def print_ascii():
    ascii_lines = [
        "\033[31m⣿⣿⣏⢴⢏⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡏⢹⣿⣀⣀⣀⣀⣀⣤⣤",
        "\033[37m⣿⣿⣿⣿⣿⣿⠿⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⠜⠻⠿⠿⠿⠿⠿⠿⠿",
        "\033[31m⣿⣿⠿⣿⣿⣷⣞⣧⣿⣿⣿⣿⣿⣿⣿⡿⠟⠋⠉⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀",
        "\033[37m⣿⣧⠾⣺⣿⣿⣿⣿⣿⣿⣿⣿⣿⣫⣤⣴⣦⣶⣾⣿⡦⠤⠀⠀⠤⠤⠤⠖⠒",
        "\033[31m⣿⣿⣿⣿⣿⡿⢋⡻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠟⢉⡀⣀⠀⢠⠃⠀⠀⠀⠀",
        "\033[37m⣿⣿⣿⣿⣿⣿⣮⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⠟⠀⡃⠀⠀⡇⡜⠉⠙⠛⠓⠒",
        "\033[31m⣿⠛⣹⢻⣿⣿⣿⣿⣿⣿⣿⡿⠿⠿⠛⠋⠉⠉⠐⠈⠙⢲⡊⡹⠳⣤⣤⡀⠀⠀",
        "\033[37m⣿⣷⣷⣿⣿⣿⣿⡿⠛⠋⠁⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡵⣁⣀⡀⠀⠉⠉⠙",
        "\033[31m⣿⣿⣿⣿⣿⡿⠋⠀⠀⠀⠀⠀⠀⠀⢀⣀⠠⠀⠀⠀⠐⠀⣇⠀⠀⠉⠓⠲⢤⡀",
        "\033[37m⣿⣿⣿⣿⡟⠀⣀⣠⣤⣶⣾⣿⣿⣿⣿⣶⣶⣦⣤⣤⣬⡀⠙⢲⡄⠀⠀⠀⠈",
        "\033[31m⣿⠟⣻⢿⣶⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠳⡄⠀⠀⠘⣦⠀",
        "\033[37m⣿⣮⣯⣾⣿⣿⡟⣩⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠋⠀⠙⠢⣄⠀⠀",
        "\033[31m⣿⣿⣿⣿⣿⣿⣿⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡆⠀⠈⠳⣄⠀⠀",
        "\033[37m⣿⣿⣿⣿⣿⣔⣻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠛⣆⠀⠀⠈⠷⢄"
    ]

    for i in range(0, len(ascii_lines), 2):
        print(ascii_lines[i])
        print(ascii_lines[i+1])
        time.sleep(1)
    print("https://github.com/pl4gu33")
    time.sleep(3)
    print("Please wait for the program to load.")
print_ascii()
print("")
print("")
print("")
def generate_password(length):
    characters = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(characters) for _ in range(length))

def create_passwords(num_passwords, app_names, lengths):
    passwords = []
    for i in range(num_passwords):
        app_name = app_names[i]
        length = lengths[i]
        password = generate_password(length)
        passwords.append((app_name, password))
    return passwords

def get_valid_input(prompt, data_type):
    while True:
        try:
            value = data_type(input(prompt))
            return value
        except ValueError:
            print("Invalid input. Please enter a valid integer.")

num_passwords = get_valid_input("Enter the number of passwords: ", int)
app_names = []
lengths = []

for i in range(num_passwords):
    app_name = input(f"Enter the name of app {i+1}: ")
    length = get_valid_input(f"Enter the password length for app {i+1}: ", int)
    app_names.append(app_name)
    lengths.append(length)

passwords = create_passwords(num_passwords, app_names, lengths)

workbook = Workbook()
worksheet = workbook.active

header_font = Font(size=16, bold=True)
content_font = Font(size=16)
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
alignment = Alignment(horizontal='center', vertical='center')

headers = ['App Name', 'Password']
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    cell = worksheet[f"{col_letter}1"]
    cell.value = header
    cell.font = header_font
    cell.border = border_style
    cell.alignment = alignment
    worksheet.column_dimensions[col_letter].width = 20

for row_num, password_pair in enumerate(passwords, 2):
    for col_num, value in enumerate(password_pair, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = value
        cell.font = content_font
        cell.border = border_style
        cell.alignment = alignment

output_path = "passwords.xlsx"
file_password = input("Enter the password to protect the .xlsx file: ")
worksheet.protection.set_password(file_password)
workbook.save(output_path)

note = "Passwords protected by the blessings of plague ;)"
note_cell = worksheet.cell(row=num_passwords + 5, column=1)
note_cell.value = note
note_cell.font = Font(size=14, bold=True, color='FF0000', underline='single')
note_cell.alignment = Alignment(horizontal='center')
worksheet.merge_cells(start_row=num_passwords+5, start_column=1, end_row=num_passwords+5, end_column=2)

output_path = "passwords.xlsx"
workbook.save(output_path)

print(f"Passwords saved successfully! Thanks for using PasswordArmour created by pl4gue.\nFile saved at: {os.path.abspath(output_path)}")

while True:
    user_choice = input("Do you want to run the program again or exit? Enter 1 to run again, 0 to exit: ")
    if user_choice == '1':
        main()
    elif user_choice == '0':
        print("Goodbye!")
        time.sleep(3)
        exit()
    else:
        print("Invalid input. Please enter either 1 or 0.")
